"""Export abgeschlossener 2.-Mahnung-Fälle als Excel im Format des Inkassobüros
und Versand per E-Mail.

Datenquelle: offene Rechnungen (im letzten Import vorhanden), deren höchste
Mahnstufe = 2 ist und die nicht als uneinbringlich markiert sind.

Zielformat: Musterdatei1.xlsx (Sheet "Beispiel MS-Import", 40 Spalten A–AN).
"""
from __future__ import annotations

import logging
import re
import sqlite3
import zipfile
from datetime import date, datetime
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.utils import formatdate
from pathlib import Path
from typing import List, Dict, Any, Optional, Tuple

import openpyxl
from openpyxl.utils import column_index_from_string

from config import InkassoConfig
from mailer import create_smtp_connection

# Feste Werte laut Absprache
SPESEN_MAHNGEBUEHR = 10.0  # 2. Mahnung: 10 € pro Fall (nicht im Rechnungsbetrag enthalten)
ZINSART = "G"              # gesetzliche Verzugszinsen -> Inkasso berechnet selbst

# Schlüsselwörter, die auf eine Firma/Institution (kein Vor-/Nachname) hindeuten
COMPANY_KEYWORDS = (
    "apotheke", "praxis", "gmbh", "mvz", " ug", " ag", " kg", "ohg", "gbr",
    "e.v", "e. v", "klinik", "zentrum", "dental", "labor", "pharma", "mbh",
    "stiftung", "verein", "institut", "gesellschaft", "co.", "kgaa",
)


# --------------------------------------------------------------------------- #
# Parsing-Helfer
# --------------------------------------------------------------------------- #
def split_name(full: str) -> Tuple[str, str, bool]:
    """(Nachname, Vorname, ist_firma). Bei Firmen steht der ganze Name im Nachnamen."""
    full = (full or "").strip()
    if not full:
        return "", "", False
    low = " " + full.lower() + " "
    is_company = (" " not in full) or any(k in low for k in COMPANY_KEYWORDS)
    if is_company:
        return full, "", True
    parts = full.split()
    return parts[-1], " ".join(parts[:-1]), False


def split_street(street: str) -> Tuple[str, str]:
    """('Neugasse 19') -> ('Neugasse', '19'). Hausnummer inkl. Zusatz (12a, 12-14)."""
    s = (street or "").strip()
    if not s:
        return "", ""
    m = re.match(r"^(.*?)[\s,]+(\d+\s*[a-zA-Z]?(?:\s*[-/]\s*\d+\s*[a-zA-Z]?)?)$", s)
    if m:
        return m.group(1).strip(), m.group(2).strip()
    return s, ""


def split_city(city: str) -> Tuple[str, str]:
    """('55232 Alzey') -> ('55232', 'Alzey')."""
    s = (city or "").strip()
    m = re.match(r"^(\d{4,5})\s+(.*)$", s)
    if m:
        return m.group(1), m.group(2).strip()
    return "", s


def _parse_date(value: Optional[str]) -> Optional[date]:
    """Akzeptiert 'YYYY-MM-DD' und 'YYYY-MM-DD HH:MM:SS' -> date."""
    if not value:
        return None
    text = str(value).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            return datetime.strptime(text[:len(fmt) + 2] if fmt.endswith("S") else text[:10], fmt).date()
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(text).date()
    except ValueError:
        return None


def _fmt_de(d: Optional[date]) -> str:
    return d.strftime("%d.%m.%Y") if d else ""


# --------------------------------------------------------------------------- #
# Datenzugriff
# --------------------------------------------------------------------------- #
def fetch_inkasso_cases(database_path: str) -> List[Dict[str, Any]]:
    """Alle offenen Fälle mit abgeschlossener 2. Mahnung (nicht uneinbringlich)."""
    with sqlite3.connect(database_path) as conn:
        conn.row_factory = sqlite3.Row
        rows = conn.execute(
            """
            WITH latest AS (SELECT MAX(snapshot_date) AS d FROM snapshots),
            last_seen AS (
                SELECT isnap.invoice_id AS invoice_id, MAX(s.snapshot_date) AS ls
                FROM invoice_snapshots isnap
                JOIN snapshots s ON s.id = isnap.snapshot_id
                GROUP BY isnap.invoice_id
            ),
            maxlvl AS (
                SELECT invoice_id, MAX(reminder_level) AS lvl
                FROM reminders GROUP BY invoice_id
            ),
            mahn2 AS (
                SELECT invoice_id, MAX(created_at) AS mahn_created
                FROM reminders WHERE reminder_level = 2 GROUP BY invoice_id
            )
            SELECT
                i.id, i.invoice_number, i.invoice_date, i.customer_name,
                i.customer_street, i.customer_city, i.amount_cents,
                cd.salutation, cd.email,
                cd.custom_name, cd.custom_street, cd.custom_city,
                mahn2.mahn_created,
                (SELECT x.file_path FROM invoice_snapshots x
                   JOIN snapshots s ON s.id = x.snapshot_id
                  WHERE x.invoice_id = i.id
                  ORDER BY s.snapshot_date DESC LIMIT 1) AS rechnung_pdf,
                (SELECT r.pdf_path FROM reminders r
                  WHERE r.invoice_id = i.id AND r.reminder_level = 2
                  ORDER BY r.created_at DESC LIMIT 1) AS mahnung2_pdf
            FROM invoices i
            JOIN last_seen ls ON ls.invoice_id = i.id
            JOIN maxlvl ml ON ml.invoice_id = i.id
            JOIN mahn2 ON mahn2.invoice_id = i.id
            LEFT JOIN customer_details cd ON cd.customer_name = i.customer_name
            WHERE ls.ls = (SELECT d FROM latest)
              AND ml.lvl = 2
              AND COALESCE(i.uncollectible, 0) = 0
            ORDER BY i.customer_name COLLATE NOCASE, i.invoice_date
            """
        ).fetchall()

    cases: List[Dict[str, Any]] = []
    for r in rows:
        name = (r["custom_name"] or r["customer_name"] or "").strip()
        street = (r["custom_street"] or r["customer_street"] or "").strip()
        city = (r["custom_city"] or r["customer_city"] or "").strip()

        nachname, vorname, _ = split_name(name)
        strasse, hausnr = split_street(street)
        plz, ort = split_city(city)
        inv_date = _parse_date(r["invoice_date"])
        mahn_date = _parse_date(r["mahn_created"])

        cases.append({
            "anrede": (r["salutation"] or "").strip(),
            "nachname": nachname,
            "vorname": vorname,
            "name_voll": name,
            "plz": plz,
            "ort": ort,
            "strasse": strasse,
            "hausnummer": hausnr,
            "strasse_mit_hnr": street,
            "email": (r["email"] or "").strip(),
            "rechnungsbetrag": (r["amount_cents"] or 0) / 100.0,
            "rechnungsdatum": inv_date,
            "mahndatum": mahn_date,
            "rechnungsnummer": str(r["invoice_number"] or "").strip(),
            "forderungsgrund": f"Unsere Rechnung Nr. {r['invoice_number']} vom {_fmt_de(inv_date)}",
            "spesen": SPESEN_MAHNGEBUEHR,
            "rechnung_pdf": (r["rechnung_pdf"] or "").strip(),
            "mahnung2_pdf": (r["mahnung2_pdf"] or "").strip(),
        })
    return cases


# --------------------------------------------------------------------------- #
# Excel-Erzeugung (auf Basis der Musterdatei)
# --------------------------------------------------------------------------- #
def build_inkasso_excel(
    cases: List[Dict[str, Any]],
    template_path: Path,
    output_path: Path,
    membership_number: str,
) -> Path:
    """Schreibt die Fälle in eine Kopie der Musterdatei (Kopfzeile bleibt erhalten)."""
    wb = openpyxl.load_workbook(template_path)
    ws = wb["Beispiel MS-Import"] if "Beispiel MS-Import" in wb.sheetnames else wb.active

    # Beispielzeilen (ab Zeile 2) entfernen, Kopfzeile behalten
    if ws.max_row >= 2:
        ws.delete_rows(2, ws.max_row - 1)

    def put(row: int, col_letter: str, value: Any) -> None:
        if value is None or value == "":
            return
        ws.cell(row=row, column=column_index_from_string(col_letter), value=value)

    for offset, c in enumerate(cases):
        row = 2 + offset
        put(row, "A", membership_number)      # Mitgliedsnummer (fest)
        # B Kundennummer / C Geschaeftszeichen -> leer
        put(row, "D", c["anrede"])            # Anrede
        put(row, "E", c["nachname"])          # Name1Nachname
        put(row, "F", c["vorname"])           # Name2Vorname
        # G Name3Geburtsname -> leer
        put(row, "H", c["name_voll"])         # Namevollstaendig
        # I Geburtsdatum -> leer
        put(row, "J", c["plz"])               # PLZ
        put(row, "K", c["ort"])               # Ort
        put(row, "L", c["strasse"])           # Strasse
        put(row, "M", c["hausnummer"])        # Hausnummer
        put(row, "N", c["strasse_mit_hnr"])   # StrassemitHNR
        # O/P/Q Telefon/Mobil/Fax -> leer
        put(row, "R", c["email"])             # Email
        # S/T/U Bank -> leer, V Zinssatz -> leer
        put(row, "W", ZINSART)                # Zinsart (gesetzlich)
        put(row, "X", c["rechnungsdatum"])    # Zinsabdatum = Rechnungsdatum
        put(row, "Y", round(c["rechnungsbetrag"], 2))  # Rechnungsbetrag
        put(row, "Z", c["rechnungsdatum"])    # Rechnungsdatum
        put(row, "AA", c["mahndatum"])        # Mahndatum = Datum 2. Mahnung
        # AB Faelligkeitsdatum -> leer
        put(row, "AC", c["rechnungsdatum"])   # Valutadatum = Rechnungsdatum
        # AD/AE Leistungszeitraum -> leer
        put(row, "AF", c["forderungsgrund"])  # Forderungsgrund
        put(row, "AG", c["rechnungsnummer"])  # Rechnungsnummer
        put(row, "AH", c["spesen"])           # Spesen = Mahngebühr (10 €)
        put(row, "AI", 0)                     # Zahlung
        put(row, "AJ", 0)                     # Gutschrift

        # Datumszellen einheitlich formatieren
        for col in ("X", "Z", "AA", "AC"):
            cell = ws.cell(row=row, column=column_index_from_string(col))
            if cell.value is not None:
                cell.number_format = "DD.MM.YYYY"
        # Betragszellen
        for col in ("Y", "AH"):
            cell = ws.cell(row=row, column=column_index_from_string(col))
            if cell.value is not None:
                cell.number_format = "#,##0.00"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    logging.info("Inkasso-Excel geschrieben: %s (%d Fälle)", output_path, len(cases))
    return output_path


def build_belege_zip(
    cases: List[Dict[str, Any]],
    data_dir: Path,
    output_path: Path,
) -> Dict[str, Any]:
    """Bündelt Rechnungs- und 2.-Mahnung-PDFs der Fälle in eine ZIP.

    Rückgabe: {'zip': Path|None, 'rechnungen': int, 'mahnungen': int, 'fehlend': [str]}.
    Mahnungen werden dedupliziert (eine Mahnung kann mehrere Rechnungen abdecken).
    """
    added_arcnames: set = set()
    missing: List[str] = []
    n_rechnungen = 0
    n_mahnungen = 0

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(output_path, "w", zipfile.ZIP_DEFLATED) as zf:
        for c in cases:
            for rel, subdir in ((c.get("rechnung_pdf"), "Rechnungen"),
                                (c.get("mahnung2_pdf"), "Mahnungen")):
                if not rel:
                    continue
                src = (data_dir / rel)
                arcname = f"{subdir}/{src.name}"
                if arcname in added_arcnames:
                    continue  # Dedupe (v.a. Sammel-Mahnungen)
                if not src.is_file():
                    missing.append(rel)
                    continue
                zf.write(src, arcname)
                added_arcnames.add(arcname)
                if subdir == "Rechnungen":
                    n_rechnungen += 1
                else:
                    n_mahnungen += 1

    if not added_arcnames:
        output_path.unlink(missing_ok=True)
        return {"zip": None, "rechnungen": 0, "mahnungen": 0, "fehlend": missing}

    logging.info(
        "Belege-ZIP geschrieben: %s (%d Rechnungen, %d Mahnungen, %d fehlend)",
        output_path, n_rechnungen, n_mahnungen, len(missing),
    )
    return {"zip": output_path, "rechnungen": n_rechnungen,
            "mahnungen": n_mahnungen, "fehlend": missing}


# --------------------------------------------------------------------------- #
# E-Mail-Versand
# --------------------------------------------------------------------------- #
def build_email_body(cfg: InkassoConfig, count: int) -> str:
    signature_lines = ["Mit freundlichen Grüßen", ""]
    if cfg.contact_name:
        signature_lines.append(cfg.contact_name)
    signature_lines.append(cfg.creditor_name)
    signature_lines.append(cfg.creditor_address)
    if cfg.contact_phone:
        signature_lines.append(f"Telefon: {cfg.contact_phone}")

    return (
        "Sehr geehrte Damen und Herren,\n\n"
        "anbei übersenden wir Ihnen im beigefügten Import-Format weitere bzw. neue "
        f"Inkasso-Fälle ({count} {'Fall' if count == 1 else 'Fälle'}), bei denen die "
        "2. Mahnung abgeschlossen ist.\n\n"
        "Ich habe die Fälle nach meinem Kenntnisstand ausgefüllt. Bitte geben Sie mir "
        "Mitteilung, wenn Sie weitere Angaben benötigen.\n\n"
        f"Unsere Mitgliedsnummer: {cfg.membership_number}\n"
        f"Für den Versand und die weitere Kommunikation nutzen Sie bitte: {cfg.smtp.user}\n\n"
        + "\n".join(signature_lines)
    )


def _attach_file(msg: MIMEMultipart, path: Path) -> None:
    """Hängt eine beliebige Datei (octet-stream) an die Nachricht an."""
    part = MIMEBase("application", "octet-stream")
    with open(path, "rb") as fh:
        part.set_payload(fh.read())
    encoders.encode_base64(part)
    part.add_header("Content-Disposition", "attachment", filename=path.name)
    msg.attach(part)


def send_inkasso_email(
    cfg: InkassoConfig,
    excel_path: Path,
    count: int,
    extra_attachments: Optional[List[Path]] = None,
) -> None:
    """Versendet die Excel-Datei (+ optionale Anhänge, z. B. Belege-ZIP) ans Inkassobüro."""
    msg = MIMEMultipart()
    msg["From"] = f"{cfg.smtp.from_name} <{cfg.smtp.user}>"
    msg["To"] = cfg.recipient
    msg["Subject"] = (
        f"Neue Inkasso-Fälle – {cfg.creditor_name} (Mitglieds-Nr. {cfg.membership_number})"
    )
    msg["Date"] = formatdate(localtime=True)
    msg.attach(MIMEText(build_email_body(cfg, count), "plain", "utf-8"))

    _attach_file(msg, excel_path)
    for extra in (extra_attachments or []):
        if extra and Path(extra).is_file():
            _attach_file(msg, Path(extra))

    connection = create_smtp_connection(cfg.smtp)
    try:
        connection.send_message(msg)
        logging.info("Inkasso-E-Mail an %s versendet (%s)", cfg.recipient, excel_path.name)
    finally:
        try:
            connection.quit()
        except Exception:
            logging.warning("SMTP-Verbindung konnte nicht sauber geschlossen werden")
